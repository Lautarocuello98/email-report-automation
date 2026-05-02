from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


FILENAME_PATTERN = re.compile(r"[^a-z0-9_-]+")
REPEATED_SEPARATOR_PATTERN = re.compile(r"[_-]{2,}")
TITLE_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
LABEL_FILL = PatternFill(fill_type="solid", fgColor="D9E2F3")
STATUS_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def load_email_template(template_path: Path) -> str:
    """Load the email template from a text file."""
    if not template_path.is_file():
        raise FileNotFoundError(f"Template file not found: {template_path}")

    return template_path.read_text(encoding="utf-8")


def _slugify(value: str) -> str:
    slug = value.strip().lower().replace(" ", "_")
    slug = FILENAME_PATTERN.sub("", slug)
    slug = REPEATED_SEPARATOR_PATTERN.sub("_", slug)
    slug = slug.strip("_-")
    return slug or "client"


def format_currency(value: float) -> str:
    return f"${value:,.2f}"


def _style_label_cell(cell) -> None:
    cell.font = Font(bold=True, color="1F1F1F")
    cell.fill = LABEL_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _style_value_cell(cell) -> None:
    cell.font = Font(color="1F1F1F")
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="left", vertical="center")


def generate_report(
    *,
    name: str,
    email: str,
    region: str,
    reporting_period: str,
    total_sales: float,
    order_count: int,
    status: str,
    output_dir: Path,
) -> Path:
    """Generate a formatted Excel report for one client."""
    output_dir.mkdir(parents=True, exist_ok=True)

    safe_name = _slugify(name)
    safe_email = _slugify(email.split("@", maxsplit=1)[0])
    report_path = output_dir / f"report_{safe_name}_{safe_email}.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Report"
    worksheet.sheet_view.showGridLines = False

    worksheet.merge_cells("A1:B1")
    worksheet["A1"] = "Client Sales Report"
    worksheet["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    worksheet["A1"].fill = TITLE_FILL
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["A1"].border = THIN_BORDER

    worksheet["A2"] = "Generated Date"
    worksheet["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    _style_label_cell(worksheet["A2"])
    _style_value_cell(worksheet["B2"])

    report_rows = [
        ("Client Name", name),
        ("Email", email),
        ("Region", region),
        ("Reporting Period", reporting_period),
        ("Total Sales", total_sales),
        ("Order Count", order_count),
        ("Status", status),
    ]

    start_row = 4
    for row_number, (label, value) in enumerate(report_rows, start=start_row):
        label_cell = worksheet[f"A{row_number}"]
        value_cell = worksheet[f"B{row_number}"]
        label_cell.value = label
        value_cell.value = value

        _style_label_cell(label_cell)
        _style_value_cell(value_cell)

        if label == "Total Sales":
            value_cell.number_format = "$#,##0.00"
            value_cell.alignment = Alignment(horizontal="right", vertical="center")
        elif label == "Order Count":
            value_cell.alignment = Alignment(horizontal="right", vertical="center")
        elif label == "Status":
            value_cell.fill = STATUS_FILL
            value_cell.font = Font(bold=True, color="1F1F1F")

    worksheet["A12"] = "Report Scope"
    worksheet["B12"] = f"This file contains report data prepared only for {name}."
    _style_label_cell(worksheet["A12"])
    _style_value_cell(worksheet["B12"])

    worksheet.column_dimensions["A"].width = 22
    worksheet.column_dimensions["B"].width = 42
    worksheet.row_dimensions[1].height = 24

    workbook.save(report_path)

    return report_path
